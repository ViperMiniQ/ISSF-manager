from openpyxl import Workbook, load_workbook
from typing import List
from openpyxl.worksheet.dimensions import ColumnDimension

workbook_extension = ".xlsx"


def create_new_workbook(name: str, path: str):
    wb = Workbook()
    wb.save(path + name + workbook_extension)


def append_rows_to_existing_workbook(name: str, path: str, rows=List[List]):
    wb = load_workbook(path + name + workbook_extension)
    ws = wb.active

    if isinstance(rows, list):
        try:
            if isinstance(rows[0], list):
                for row in rows:
                    ws.append(row)
            else:
                ws.append(rows)
        except IndexError:
            pass

    wb.save(path + name + workbook_extension)


def export_to_excel_workbook(sheet_name: str, data: List[List], path: str, filename: str, headers: List = None):
    wb = Workbook()
    ws = wb.active
    ColumnDimension(ws, bestFit=True)
    ws.title = sheet_name
    row = 1
    column = 1

    for header in headers:
        ws.cell(row=row, column=column, value=header)
        column += 1

    row += 1
    column = 1

    for row_data in data:
        for cell_value in row_data:
            ws.cell(row=row, column=column, value=cell_value)
            column += 1
        row += 1
        column = 1

    wb.save(path + filename + workbook_extension)